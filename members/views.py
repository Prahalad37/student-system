import openpyxl
import xlwt
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.template import loader
from django.template.loader import get_template
from django.db.models import Sum
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from datetime import date
from xhtml2pdf import pisa

# ✅ Import Models
from .models import Member, Attendance, StudyMaterial, ExamScore, Notice, Expense, ClassRoom

# ==========================================
#               DASHBOARD
# ==========================================
@login_required
def index(request):
    mymembers = Member.objects.all()
    total_students = Member.objects.count()
    
    # Notices (Newest first)
    notices = Notice.objects.all().order_by('-id')

    # Revenue & Expense Logic
    revenue_data = Member.objects.aggregate(Sum('fee_paid'))
    total_revenue = revenue_data['fee_paid__sum'] if revenue_data['fee_paid__sum'] else 0

    expense_data = Expense.objects.aggregate(Sum('amount'))
    total_expense = expense_data['amount__sum'] if expense_data['amount__sum'] else 0

    net_profit = total_revenue - total_expense

    # Attendance Logic
    total_records = Attendance.objects.count()
    present_records = Attendance.objects.filter(status='Present').count()
    attendance_percent = round((present_records / total_records) * 100) if total_records > 0 else 0

    context = {
        'mymembers': mymembers,
        'total_students': total_students,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_profit': net_profit,
        'attendance_percent': attendance_percent,
        'notices': notices,
    }
    return render(request, 'index.html', context)

# ==========================================
#           STUDENT MANAGEMENT
# ==========================================
@login_required
def all_students(request):
    """View All Students List"""
    mymembers = Member.objects.all().order_by('firstname')
    return render(request, 'all_students.html', {'mymembers': mymembers})

@login_required
def student_profile(request, id):
    """View Single Student Profile"""
    student = get_object_or_404(Member, id=id)
    
    # Sort exams by ID
    exams = ExamScore.objects.filter(student=student).order_by('-id')
    
    attendance_list = Attendance.objects.filter(student=student).order_by('-date')
    
    total_days = attendance_list.count()
    present_days = attendance_list.filter(status='Present').count()
    
    if total_days > 0:
        attendance_percent = round((present_days / total_days) * 100, 1)
    else:
        attendance_percent = 0
        
    context = {
        'student': student,
        'exams': exams,
        'attendance_list': attendance_list,
        'attendance_percent': attendance_percent,
        'present_days': present_days,
        'total_days': total_days
    }
    return render(request, 'student_profile.html', context)

@login_required
def add(request):
    """Render Add Form"""
    template = loader.get_template('add.html')
    return HttpResponse(template.render({}, request))

@login_required
def addrecord(request):
    """Save New Student"""
    if request.method == "POST":
        # 1. Basic Fields
        first = request.POST['first']
        last = request.POST['last']
        father = request.POST.get('father_name')
        mobile = request.POST.get('mobile_number')
        email = request.POST.get('email')
        
        # 2. Official & Personal
        admission_no = request.POST.get('admission_no')
        roll_number = request.POST.get('roll_number')
        address = request.POST.get('address')
        joined_date = request.POST.get('joined_date') or None
        dob = request.POST.get('dob') or None
        gender = request.POST.get('gender') or 'Male'

        # 3. Class Logic
        class_name = request.POST.get('student_class')
        section_name = request.POST.get('section') or 'A'
        
        class_obj = None
        if class_name:
            class_obj, created = ClassRoom.objects.get_or_create(
                name=class_name, 
                section=section_name
            )

        # 4. Other Fields
        house = request.POST.get('house_team')
        stream = request.POST.get('preferred_stream')
        blood = request.POST.get('blood_group')
        medical = request.POST.get('medical_conditions')
        transport = request.POST.get('transport_mode')
        route = request.POST.get('route_name')
        
        # 5. Fees & Image
        fee_total = request.POST.get('fee_total', 0)
        fee_paid = request.POST.get('fee_paid', 0)
        
        img = request.FILES.get('file')
            
        # 6. Save
        member = Member(
            firstname=first, lastname=last, father_name=father,       
            mobile_number=mobile, email=email, admission_no=admission_no,
            joined_date=joined_date, roll_number=roll_number, gender=gender,
            dob=dob, address=address, student_class=class_obj,
            house_team=house, preferred_stream=stream, blood_group=blood,
            medical_conditions=medical, transport_mode=transport, route_name=route,
            fee_total=fee_total, fee_paid=fee_paid, profile_pic=img
        )
        member.save()
        return HttpResponseRedirect(reverse('index'))
    
    return redirect('add')

@login_required
def delete(request, id):
    member = get_object_or_404(Member, id=id)
    member.delete()
    return HttpResponseRedirect(reverse('index'))

@login_required
def update(request, id):
    mymember = get_object_or_404(Member, id=id)
    template = loader.get_template('update.html')
    return HttpResponse(template.render({'mymember': mymember}, request))

@login_required
def updaterecord(request, id):
    if request.method == "POST":
        member = get_object_or_404(Member, id=id)

        # Update Fields
        member.firstname = request.POST['first']
        member.lastname = request.POST['last']
        member.father_name = request.POST.get('father_name')
        member.mobile_number = request.POST.get('mobile_number')
        member.email = request.POST.get('email')
        member.address = request.POST.get('address')
        member.admission_no = request.POST.get('admission_no')
        member.roll_number = request.POST.get('roll_number')
        member.gender = request.POST.get('gender')
        
        if request.POST.get('joined_date'): member.joined_date = request.POST.get('joined_date')
        if request.POST.get('dob'): member.dob = request.POST.get('dob')

        # Class Update
        class_name = request.POST.get('student_class')
        section_name = request.POST.get('section') or 'A'
        if class_name:
            class_obj, created = ClassRoom.objects.get_or_create(name=class_name, section=section_name)
            member.student_class = class_obj

        # Other Details
        member.house_team = request.POST.get('house_team')
        member.preferred_stream = request.POST.get('preferred_stream')
        member.blood_group = request.POST.get('blood_group')
        member.medical_conditions = request.POST.get('medical_conditions')
        member.transport_mode = request.POST.get('transport_mode')
        member.route_name = request.POST.get('route_name')
        member.fee_total = request.POST.get('fee_total', 0)
        member.fee_paid = request.POST.get('fee_paid', 0)

        if request.FILES.get('file'):
            member.profile_pic = request.FILES['file']

        member.save()
        return HttpResponseRedirect(reverse('index'))
        
    return redirect('index')

# ==========================================
#          ATTENDANCE SYSTEM
# ==========================================
@login_required
def attendance(request):
    if request.method == "POST":
        selected_date = request.POST.get('date')
        all_students = Member.objects.all()
        
        for student in all_students:
            status_value = request.POST.get(f'status_{student.id}')
            if status_value:
                Attendance.objects.update_or_create(
                    student=student,
                    date=selected_date,
                    defaults={'status': status_value}
                )
        return redirect('attendance_records')

    mymembers = Member.objects.all()
    return render(request, 'attendance.html', {'mymembers': mymembers})

@login_required
def attendance_records(request):
    records = Attendance.objects.all().order_by('-date', 'student__firstname')
    return render(request, 'attendance_records.html', {'records': records})

# ==========================================
#           ACADEMICS (Library & Exams)
# ==========================================
@login_required
def library(request):
    if request.method == "POST":
        title = request.POST.get('title')
        subject = request.POST.get('subject')
        class_name = request.POST.get('class_name')
        video_link = request.POST.get('video_link')
        
        pdf_file = None
        if request.FILES.get('pdf_file'):
            pdf = request.FILES['pdf_file']
            fs = FileSystemStorage()
            filename = fs.save(pdf.name, pdf)
            pdf_file = fs.url(filename)

        StudyMaterial.objects.create(
            title=title, subject=subject, class_name=class_name,
            video_link=video_link, pdf_file=pdf_file
        )
        return redirect('library')

    materials = StudyMaterial.objects.all().order_by('-id')
    return render(request, 'library.html', {'materials': materials})

@login_required
def add_marks(request):
    if request.method == "POST":
        student_id = request.POST.get('student_id')
        exam_name = request.POST.get('exam_name')
        student = get_object_or_404(Member, id=student_id)
        
        ExamScore.objects.create(
            student=student, exam_name=exam_name,
            maths=request.POST.get('maths'), physics=request.POST.get('physics'),
            chemistry=request.POST.get('chemistry'), english=request.POST.get('english'),
            computer=request.POST.get('computer')
        )
        return redirect('report_card')

    students = Member.objects.all()
    return render(request, 'add_marks.html', {'students': students})

@login_required
def report_card(request):
    scores = ExamScore.objects.all().order_by('-id')
    return render(request, 'report_card.html', {'scores': scores})

@login_required
def marksheet_pdf(request, id):
    score = get_object_or_404(ExamScore, id=id)
    
    total_obtained = score.maths + score.physics + score.chemistry + score.english + score.computer
    total_max = 500
    percentage = round((total_obtained / total_max) * 100, 2)
    result_status = "PASS" if percentage >= 33 else "FAIL"
    
    context = {
        'score': score,
        'student': score.student,
        'total_obtained': total_obtained,
        'total_max': total_max,
        'percentage': percentage,
        'result_status': result_status,
        'date': date.today(),
    }
    
    template = get_template('marksheet_pdf.html')
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="ReportCard_{score.student.firstname}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse(f'PDF Error: <pre>{html}</pre>')
    return response

# ==========================================
#           FINANCE (Receipts & Expenses)
# ==========================================
@login_required
def receipt_pdf(request, id):
    student = get_object_or_404(Member, id=id)
    context = {
        'student': student,
        'date': date.today(),
        'receipt_no': f"REC-{student.id}-{date.today().strftime('%m%d')}"
    }
    
    template = get_template('receipt_pdf.html')
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Receipt_{student.firstname}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse(f'PDF Error: <pre>{html}</pre>')
    return response

@login_required
def add_expense(request):
    if request.method == "POST":
        description = request.POST.get('description')
        amount = request.POST.get('amount')
        Expense.objects.create(description=description, amount=amount)
        return redirect('add_expense')

    expenses = Expense.objects.all().order_by('-id')
    return render(request, 'add_expense.html', {'expenses': expenses})

# ==========================================
#           NOTICES & ID CARDS
# ==========================================
@login_required
def add_notice(request):
    if request.method == "POST":
        title = request.POST.get('title')
        message = request.POST.get('message')
        Notice.objects.create(title=title, message=message)
    return redirect('index')

@login_required
def delete_notice(request, id):
    notice = get_object_or_404(Notice, id=id)
    notice.delete()
    return redirect('index')

@login_required
def id_card(request, id):
    student = get_object_or_404(Member, id=id)
    return render(request, 'id_card.html', {'student': student})

# ==========================================
#           IMPORT / EXPORT (Excel)
# ==========================================
@login_required
def import_students(request):
    if request.method == "POST" and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
            
            # Start from Row 2
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                fname = row[0]
                lname = row[1]
                class_name = str(row[2]) if row[2] else None
                section = str(row[3]) if row[3] else 'A'
                roll_no = row[4]
                mobile = row[5]
                fee_total = row[6] if row[6] else 0
                fee_paid = row[7] if row[7] else 0
                
                if fname:
                    # Class Setup
                    class_obj = None
                    if class_name:
                        class_obj, created = ClassRoom.objects.get_or_create(
                            name=class_name, section=section
                        )
                    
                    # Crash-Proof Logic: Search Existing
                    target_member = None
                    if roll_no and class_obj:
                        candidates = Member.objects.filter(roll_number=roll_no, student_class=class_obj)
                        if candidates.exists(): target_member = candidates.first()
                    
                    if not target_member and class_obj:
                        candidates = Member.objects.filter(firstname=fname, lastname=lname, student_class=class_obj)
                        if candidates.exists(): target_member = candidates.first()

                    # Update or Create
                    if target_member:
                        target_member.firstname = fname
                        target_member.lastname = lname
                        target_member.roll_number = roll_no
                        target_member.mobile_number = mobile
                        target_member.fee_total = fee_total
                        target_member.fee_paid = fee_paid
                        target_member.save()
                    else:
                        Member.objects.create(
                            firstname=fname, lastname=lname, student_class=class_obj,
                            roll_number=roll_no, mobile_number=mobile,
                            fee_total=fee_total, fee_paid=fee_paid
                        )
            return redirect('index')
        except Exception as e:
            return HttpResponse(f"Error in Excel File: {e}")

    return render(request, 'import_students.html')

@login_required
def export_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Student_List.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Students')

    # Header
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['ID', 'First Name', 'Last Name', 'Class', 'Total Fees', 'Paid Fees', 'Due Amount']

    for col_num, col_name in enumerate(columns):
        ws.write(0, col_num, col_name, font_style)

    # Body
    font_style = xlwt.XFStyle()
    members = Member.objects.all()
    
    for row_num, member in enumerate(members, start=1):
        class_str = str(member.student_class) if member.student_class else "N/A"
        due_amount = member.fee_total - member.fee_paid
        
        row = [
            member.id, member.firstname, member.lastname, class_str,
            member.fee_total, member.fee_paid, due_amount
        ]
        for col_num, cell_value in enumerate(row):
            ws.write(row_num, col_num, cell_value, font_style)

    wb.save(response)
    return response